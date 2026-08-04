#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
"""
设备信息配置模块
支持手动输入和Excel批量导入
支持IPv4和IPv6地址
"""

from typing import List, Dict, Optional
import os
import json
import sys
from openpyxl import Workbook, load_workbook

# 添加项目路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.ipv6_utils import IPv6Utils, IPVersion, IPv6AddressValidator
from utils.password_crypto import decrypt_password, encrypt_password, is_encrypted_password


SUPPORTED_DEVICE_BRANDS = ("h3c", "huawei")


def normalize_supported_brand(brand: str, default: str = "h3c") -> str:
    """Normalize user-facing brands and reject brands without verified support."""
    value = str(brand or default).strip().lower()
    aliases = {
        "h3c": "h3c",
        "comware": "h3c",
        "huawei": "huawei",
        "vrp": "huawei",
    }
    normalized = aliases.get(value)
    if normalized:
        return normalized
    raise ValueError(
        "unsupported brand; currently supported brands are H3C and Huawei"
    )


class DeviceInfo:
    """设备信息类（支持IPv4和IPv6）"""
    
    def __init__(self, brand: str = '', ip: str = '', port: int = 22,
                 username: str = '', password: str = '', name: str = '',
                 group: str = '', tags: str = '', auth_method: str = 'password',
                 private_key_path: str = '', private_key_passphrase: str = '',
                 host_key_policy: str = 'tofu'):
        self.brand = brand
        self.ip = ip
        self.port = port
        self.username = username
        self.password = password
        self.name = name or f"{brand}_{ip}"
        self.group = str(group or '').strip()
        self.tags = self._normalize_tags(tags)
        self.auth_method = (
            'key' if str(auth_method or '').strip().lower() == 'key' else 'password'
        )
        self.private_key_path = str(private_key_path or '').strip()
        self.private_key_passphrase = str(private_key_passphrase or '')
        self.host_key_policy = str(host_key_policy or 'tofu').strip().lower()
        
        # IPv6相关属性
        self.ip_version = IPVersion.UNKNOWN
        self.ipv6_validator = IPv6AddressValidator()
        
        # 验证IP地址并记录版本
        if ip:
            self.ip_version = IPv6Utils.get_ip_version(ip)
    
    def validate_ip_address(self) -> tuple:
        """
        验证IP地址
        
        Returns:
            tuple: (是否有效, 错误信息)
        """
        if not self.ip:
            return False, "IP地址不能为空"
        
        return self.ipv6_validator.validate_for_ssh(self.ip)
    
    def get_ip_info(self) -> dict:
        """
        获取IP地址的详细信息
        
        Returns:
            dict: IP地址信息
        """
        return IPv6Utils.get_ip_address_info(self.ip)
    
    def get_display_address(self) -> str:
        """
        获取用于显示的地址格式
        
        Returns:
            str: 显示格式的地址
        """
        return IPv6Utils.format_ipv6_for_display(self.ip)
    
    @staticmethod
    def _normalize_tags(tags) -> str:
        values = tags if isinstance(tags, (list, tuple, set)) else (
            str(tags or '').replace('，', ',').split(',')
        )
        cleaned = []
        for value in values:
            tag = str(value or '').strip()
            if tag and tag not in cleaned:
                cleaned.append(tag)
        return ','.join(cleaned)

    def to_dict(self, include_secrets: bool = True) -> Dict:
        """转换为字典"""
        return {
            'name': self.name,
            'brand': self.brand,
            'ip': self.ip,
            'port': self.port,
            'username': self.username,
            'password': self.password if include_secrets else '',
            'group': self.group,
            'tags': self.tags,
            'auth_method': self.auth_method,
            'private_key_path': self.private_key_path,
            'private_key_passphrase': (
                self.private_key_passphrase if include_secrets else ''
            ),
            'host_key_policy': self.host_key_policy,
            'ip_version': self.ip_version.value if self.ip_version else 0,
            'ip_version_name': 'IPv6' if self.ip_version == IPVersion.IPv6 else 'IPv4' if self.ip_version == IPVersion.IPv4 else 'Unknown',
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'DeviceInfo':
        """从字典创建"""
        device = cls(
            brand=data.get('brand', ''),
            ip=data.get('ip', ''),
            port=data.get('port', 22),
            username=data.get('username', ''),
            password=data.get('password', ''),
            name=data.get('name', ''),
            group=data.get('group', ''),
            tags=data.get('tags', ''),
            auth_method=data.get('auth_method', 'password'),
            private_key_path=data.get('private_key_path', ''),
            private_key_passphrase=data.get('private_key_passphrase', ''),
            host_key_policy=data.get('host_key_policy', 'tofu'),
        )
        
        # 恢复IP版本信息
        if 'ip_version' in data:
            version = data['ip_version']
            if version == 6:
                device.ip_version = IPVersion.IPv6
            elif version == 4:
                device.ip_version = IPVersion.IPv4
        
        return device
    
    def __str__(self) -> str:
        ip_display = self.get_display_address()
        version_info = f" [{self.ip_version.value}]" if self.ip_version != IPVersion.UNKNOWN else ""
        return f"DeviceInfo({self.name}, {ip_display}:{self.port}{version_info})"
    
    def __repr__(self) -> str:
        return self.__str__()


class DeviceConfigManager:
    """设备配置管理器"""
    
    def __init__(self):
        self.devices: List[DeviceInfo] = []
        self.last_import_skipped_count = 0
        self.last_import_skipped = []

    @staticmethod
    def _device_key(ip: str, port: int = 22) -> tuple:
        ip = str(ip or '').strip()
        version = IPv6Utils.get_ip_version(ip)
        if version == IPVersion.IPv6:
            ip = IPv6Utils.remove_ipv6_scope_id(ip)
            ip = IPv6Utils.normalize_ipv6(ip).lower()
        else:
            ip = ip.lower()
        return ip, int(port or 22)

    def has_device(self, ip: str, port: int = 22) -> bool:
        key = self._device_key(ip, port)
        return any(self._device_key(device.ip, device.port) == key for device in self.devices)
    
    def add_device(self, device: DeviceInfo):
        """Add one device, skipping duplicate IP+port entries."""
        if self.has_device(device.ip, device.port):
            return False
        self.devices.append(device)
        return True
    
    def add_device_manual(self, brand: str, ip: str, port: int = 22,
                          username: str = '', password: str = '', name: str = '',
                          **kwargs):
        """手动添加设备"""
        brand = normalize_supported_brand(brand)
        device = DeviceInfo(brand, ip, port, username, password, name, **kwargs)
        self.add_device(device)
    
    @staticmethod
    def _clean_excel_value(value, default: str = '') -> str:
        """Return a safe string for Excel cells, treating NaN/blank as default."""
        if value is None:
            return default
        value = str(value).strip()
        if value.lower() in ('nan', 'none'):
            return default
        return value

    @staticmethod
    def _clean_excel_port(value, default: int = 22) -> int:
        """Normalize an Excel port cell and validate the TCP port range."""
        if value is None or str(value).strip() == '':
            return default
        try:
            port = int(float(value))
        except (TypeError, ValueError):
            raise ValueError('port must be a number between 1 and 65535')
        if not 1 <= port <= 65535:
            raise ValueError('port must be between 1 and 65535')
        return port

    @classmethod
    def _device_from_mapping(cls, data: Dict) -> DeviceInfo:
        """Create a validated DeviceInfo from a JSON-like mapping."""
        if not isinstance(data, dict):
            raise ValueError('device item must be an object')

        brand = normalize_supported_brand(
            cls._clean_excel_value(data.get('brand'), 'h3c')
        )
        ip = cls._clean_excel_value(data.get('ip'))
        port = cls._clean_excel_port(data.get('port'), 22)
        username = cls._clean_excel_value(data.get('username'))
        password = cls._clean_excel_value(data.get('password'))
        name = cls._clean_excel_value(data.get('name'))
        group = cls._clean_excel_value(data.get('group'))
        tags = cls._clean_excel_value(data.get('tags'))
        auth_method = cls._clean_excel_value(
            data.get('auth_method'), 'password'
        ).lower()
        private_key_path = cls._clean_excel_value(data.get('private_key_path'))
        private_key_passphrase = cls._clean_excel_value(
            data.get('private_key_passphrase')
        )
        host_key_policy = cls._clean_excel_value(
            data.get('host_key_policy'), 'tofu'
        ).lower()

        if not ip:
            raise ValueError('ip is required')
        if not username:
            raise ValueError('username is required')
        if auth_method not in ('password', 'key'):
            raise ValueError('auth_method must be password or key')
        if auth_method == 'password' and not password:
            raise ValueError('password is required')
        if auth_method == 'key' and not private_key_path:
            raise ValueError('private_key_path is required for key authentication')

        device = DeviceInfo(
            brand, ip, port, username, password, name,
            group=group, tags=tags, auth_method=auth_method,
            private_key_path=private_key_path,
            private_key_passphrase=private_key_passphrase,
            host_key_policy=host_key_policy,
        )
        is_valid, error_msg = device.validate_ip_address()
        if not is_valid:
            raise ValueError(f'invalid IP address: {error_msg}')
        return device

    @staticmethod
    def inspect_excel_password_mode(file_path: str) -> str:
        """Return encryption mode for password and private-key passphrase cells."""
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or '').strip().lower() for value in next(rows, [])]
            indexes = [
                headers.index(name)
                for name in ('password', 'private_key_passphrase')
                if name in headers
            ]
            if not indexes:
                return 'none'
            encrypted = plain = False
            for row in rows:
                for column in indexes:
                    value = row[column] if column < len(row) else None
                    if value is None or str(value).strip() == '':
                        continue
                    if is_encrypted_password(str(value)):
                        encrypted = True
                    else:
                        plain = True
            if encrypted and plain:
                return 'mixed'
            if not encrypted and not plain:
                return 'none'
            return 'encrypted' if encrypted else 'plain'
        finally:
            workbook.close()

    @staticmethod
    def encrypt_excel_passwords(source_path: str, target_path: str, master_password: str) -> int:
        """Copy an Excel file while encrypting every non-empty password cell."""
        if len(master_password or '') < 8:
            raise ValueError('主密码至少需要 8 个字符')
        workbook = load_workbook(source_path)
        try:
            sheet = workbook.active
            headers = [str(cell.value or '').strip().lower() for cell in sheet[1]]
            protected_columns = [
                headers.index(name) + 1
                for name in ('password', 'private_key_passphrase')
                if name in headers
            ]
            if not protected_columns:
                raise ValueError('Excel 缺少 password 或 private_key_passphrase 列')
            encrypted_count = 0
            for column in protected_columns:
                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=column)
                    value = str(cell.value or '').strip()
                    if not value or is_encrypted_password(value):
                        continue
                    cell.value = encrypt_password(value, master_password)
                    encrypted_count += 1
            if encrypted_count == 0:
                raise ValueError('没有找到可加密的明文密码')
            workbook.save(target_path)
            return encrypted_count
        finally:
            workbook.close()

    def import_from_excel(self, file_path: str, master_password: str = '') -> tuple:
        """Import devices from Excel with authentication and IP validation."""
        success_count = 0
        error_count = 0
        errors = []
        self.last_import_skipped_count = 0
        self.last_import_skipped = []
        workbook = None
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            if not rows:
                return 0, 0, ["Excel file is empty"]

            headers = [self._clean_excel_value(value).lower() for value in rows[0]]
            data_rows = rows[1:]
            missing = [name for name in ('ip', 'username') if name not in headers]
            if missing:
                return 0, len(data_rows), [
                    f"Excel file missing required columns: {missing}"
                ]
            header_index = {name: index for index, name in enumerate(headers) if name}

            for index, values in enumerate(data_rows):
                row_no = index + 2
                row = {
                    name: values[column] if column < len(values) else None
                    for name, column in header_index.items()
                }
                try:
                    brand = normalize_supported_brand(
                        self._clean_excel_value(row.get('brand'), 'h3c')
                    )
                    ip = self._clean_excel_value(row.get('ip'))
                    port = self._clean_excel_port(row.get('port'), 22)
                    username = self._clean_excel_value(row.get('username'))
                    password = self._clean_excel_value(row.get('password'))
                    if is_encrypted_password(password):
                        password = decrypt_password(password, master_password)
                    name = self._clean_excel_value(row.get('name'))
                    group = self._clean_excel_value(row.get('group'))
                    tags = self._clean_excel_value(row.get('tags'))
                    auth_method = self._clean_excel_value(
                        row.get('auth_method'), 'password'
                    ).lower()
                    private_key_path = self._clean_excel_value(
                        row.get('private_key_path')
                    )
                    if private_key_path and not os.path.isabs(private_key_path):
                        private_key_path = os.path.abspath(os.path.join(
                            os.path.dirname(file_path), private_key_path
                        ))
                    private_key_passphrase = self._clean_excel_value(
                        row.get('private_key_passphrase')
                    )
                    if is_encrypted_password(private_key_passphrase):
                        private_key_passphrase = decrypt_password(
                            private_key_passphrase, master_password
                        )
                    host_key_policy = self._clean_excel_value(
                        row.get('host_key_policy'), 'tofu'
                    ).lower()

                    if not ip:
                        raise ValueError('ip is required')
                    if not username:
                        raise ValueError('username is required')
                    if auth_method not in ('password', 'key'):
                        raise ValueError('auth_method must be password or key')
                    if auth_method == 'password' and not password:
                        raise ValueError('password is required')
                    if auth_method == 'key' and not private_key_path:
                        raise ValueError(
                            'private_key_path is required for key authentication'
                        )

                    device = DeviceInfo(
                        brand, ip, port, username, password, name,
                        group=group, tags=tags, auth_method=auth_method,
                        private_key_path=private_key_path,
                        private_key_passphrase=private_key_passphrase,
                        host_key_policy=host_key_policy,
                    )
                    is_valid, error_msg = device.validate_ip_address()
                    if not is_valid:
                        raise ValueError(f'invalid IP address: {error_msg}')
                    if not self.add_device(device):
                        self.last_import_skipped_count += 1
                        self.last_import_skipped.append(
                            f"Row {row_no}: duplicate {ip}:{port}"
                        )
                        continue
                    success_count += 1
                except Exception as exc:
                    error_count += 1
                    errors.append(f"Row {row_no} import failed: {exc}")
        except Exception as exc:
            errors.append(f"Excel file read failed: {exc}")
            error_count = 1
        finally:
            if workbook is not None:
                workbook.close()
        return success_count, error_count, errors

    def export_to_excel(self, file_path: str, include_password: bool = False,
                        master_password: str = '') -> bool:
        """
        导出设备信息到Excel
        
        Args:
            file_path: 输出文件路径
            
        Returns:
            bool: 是否成功
        """
        try:
            workbook = Workbook()
            sheet = workbook.active
            headers = [
                'name', 'group', 'tags', 'brand', 'ip', 'port', 'username',
                'auth_method', 'password', 'private_key_path',
                'private_key_passphrase', 'host_key_policy',
                'ip_version', 'ip_version_name',
            ]
            sheet.append(headers)
            for device in self.devices:
                data = device.to_dict()
                if include_password and master_password:
                    data['password'] = encrypt_password(data['password'], master_password)
                    if data.get('private_key_passphrase'):
                        data['private_key_passphrase'] = encrypt_password(
                            data['private_key_passphrase'], master_password
                        )
                else:
                    data['password'] = ''
                    data['private_key_passphrase'] = ''
                sheet.append([data.get(header, '') for header in headers])
            workbook.save(file_path)
            return True
        except Exception as e:
            print(f"导出失败: {str(e)}")
            return False
    
    def save_to_json(self, file_path: str) -> bool:
        """保存到JSON文件"""
        try:
            data = [device.to_dict() for device in self.devices]
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存失败: {str(e)}")
            return False
    
    def load_from_json(self, file_path: str) -> bool:
        """从JSON文件加载"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if not isinstance(data, list):
                raise ValueError('JSON root must be a device list')

            loaded_devices = []
            skipped_errors = []
            seen_keys = set()
            self.last_import_skipped_count = 0
            self.last_import_skipped = []
            for index, item in enumerate(data, start=1):
                try:
                    device = self._device_from_mapping(item)
                    key = self._device_key(device.ip, device.port)
                    if key in seen_keys:
                        self.last_import_skipped_count += 1
                        self.last_import_skipped.append(f"Item {index}: duplicate {device.ip}:{device.port}")
                        continue
                    seen_keys.add(key)
                    loaded_devices.append(device)
                except Exception as e:
                    skipped_errors.append(f"Item {index}: {str(e)}")

            if data and not loaded_devices:
                raise ValueError('; '.join(skipped_errors) or 'no valid devices')

            if skipped_errors:
                print(f"JSON load skipped invalid devices: {'; '.join(skipped_errors)}")

            self.devices = loaded_devices
            return True
        except Exception as e:
            print(f"加载失败: {str(e)}")
            return False
    
    def get_devices(self) -> List[DeviceInfo]:
        """获取所有设备"""
        return self.devices
    
    def clear_devices(self):
        """清空设备列表"""
        self.devices.clear()
    
    def remove_device(self, index: int):
        """移除指定设备"""
        if 0 <= index < len(self.devices):
            self.devices.pop(index)
    
    def get_device_count(self) -> int:
        """获取设备数量"""
        return len(self.devices)
    
    def create_template_excel(self, file_path: str) -> bool:
        """创建设备信息模板Excel"""
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "devices"
            sheet.append([
                'name', 'group', 'tags', 'brand', 'ip', 'port', 'username',
                'auth_method', 'password', 'private_key_path',
                'private_key_passphrase', 'host_key_policy',
            ])
            sheet.append([
                'SW1', '核心交换机', '机房A,核心', 'h3c', '192.168.1.1',
                22, 'admin', 'password', 'password1', '', '', 'tofu',
            ])
            sheet.append([
                'SW2', '接入交换机', '机房A,接入', 'h3c', '192.168.1.2',
                22, 'admin', 'key', '', '.ssh/id_ed25519', '', 'strict',
            ])
            workbook.save(file_path)
            return True
        except Exception as e:
            print(f"创建模板失败: {str(e)}")
            return False
