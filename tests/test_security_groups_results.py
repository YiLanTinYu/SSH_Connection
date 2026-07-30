import csv
import json
import os
import socket
import threading
from datetime import datetime, timedelta
from pathlib import Path

import paramiko
from openpyxl import Workbook, load_workbook

from config.device_config import DeviceConfigManager, DeviceInfo
from config.ssh_security import (
    HOST_KEY_STRICT,
    HOST_KEY_TOFU,
    build_connect_kwargs,
    configure_host_key_policy,
    persist_host_keys,
)
from core.ssh_manager_simple import SSHConnection
from utils.result_export import (
    export_results_csv,
    export_results_json,
    export_results_xlsx,
)


class FakeSSHClient:
    def __init__(self):
        self.loaded_system = False
        self.loaded = []
        self.saved = []
        self.policy = None

    def load_system_host_keys(self):
        self.loaded_system = True

    def load_host_keys(self, path):
        self.loaded.append(path)

    def set_missing_host_key_policy(self, policy):
        self.policy = policy

    def save_host_keys(self, path):
        self.saved.append(path)


def test_device_metadata_and_public_result_do_not_expose_secrets():
    device = DeviceInfo(
        "h3c", "192.0.2.10", 22, "admin", "secret", "SW1",
        group="核心", tags="机房A，核心,核心", auth_method="key",
        private_key_path="id_ed25519", private_key_passphrase="key-secret",
        host_key_policy="strict",
    )
    assert device.group == "核心"
    assert device.tags == "机房A,核心"

    connection = SSHConnection(device)
    result = connection.get_connection_info()
    assert result["device_info"]["password"] == ""
    assert result["device_info"]["private_key_passphrase"] == ""
    assert result["device_info"]["private_key_path"] == "id_ed25519"


def test_private_key_connect_arguments(tmp_path):
    key_path = tmp_path / "id_ed25519"
    key_path.write_text("test key placeholder", encoding="ascii")
    device = DeviceInfo(
        "h3c", "192.0.2.10", 22, "admin", "", "SW1",
        auth_method="key", private_key_path=str(key_path),
        private_key_passphrase="passphrase",
    )
    kwargs = build_connect_kwargs(device, device.ip)
    assert kwargs["key_filename"] == str(key_path)
    assert kwargs["passphrase"] == "passphrase"
    assert "password" not in kwargs
    assert kwargs["allow_agent"] is False
    assert kwargs["look_for_keys"] is False


def test_host_key_policy_loads_and_persists_tofu(tmp_path):
    known_hosts = tmp_path / "known_hosts"
    known_hosts.write_text("", encoding="ascii")
    client = FakeSSHClient()

    path = configure_host_key_policy(client, HOST_KEY_TOFU, str(known_hosts))
    persist_host_keys(client, HOST_KEY_TOFU, path)

    assert client.loaded_system is True
    assert client.loaded == [str(known_hosts)]
    assert client.saved == [str(known_hosts)]
    assert client.policy.__class__.__name__ == "AutoAddPolicy"


def test_strict_host_key_policy_never_persists(tmp_path):
    client = FakeSSHClient()
    path = str(tmp_path / "missing_known_hosts")
    configure_host_key_policy(client, HOST_KEY_STRICT, path)
    persist_host_keys(client, HOST_KEY_STRICT, path)
    assert client.policy.__class__.__name__ == "RejectPolicy"
    assert client.saved == []


def test_excel_import_supports_key_auth_group_and_tags(tmp_path):
    key_path = tmp_path / "switch.key"
    key_path.write_text("placeholder", encoding="ascii")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "name", "group", "tags", "brand", "ip", "port", "username",
        "auth_method", "password", "private_key_path",
        "private_key_passphrase", "host_key_policy",
    ])
    sheet.append([
        "SW1", "核心", "机房A,核心", "h3c", "192.0.2.10", 22, "admin",
        "key", "", key_path.name, "", "strict",
    ])
    source = tmp_path / "devices.xlsx"
    workbook.save(source)

    manager = DeviceConfigManager()
    success, failure, errors = manager.import_from_excel(str(source))
    assert (success, failure, errors) == (1, 0, [])
    device = manager.get_devices()[0]
    assert device.group == "核心"
    assert device.auth_method == "key"
    assert device.private_key_path == str(key_path)
    assert device.host_key_policy == "strict"


def _sample_results():
    return [{
        "device_info": {
            "name": "SW1", "group": "核心", "tags": "机房A",
            "ip": "192.0.2.10", "brand": "h3c",
        },
        "task_success": True,
        "brand_detected": "h3c",
        "model_detected": "S6520",
        "error_message": "",
        "started_at": "2026-07-27T10:00:00",
        "finished_at": "2026-07-27T10:00:02",
        "duration_seconds": 2.0,
        "connection_duration_seconds": 1.2,
        "operation_duration_seconds": 0.8,
        "command_results": [{
            "command": "display version",
            "output": "H3C Comware Software",
            "timestamp": "2026-07-27T10:00:01",
            "duration_seconds": 0.8,
        }],
    }]


def test_result_exports_include_summary_and_command_details(tmp_path):
    results = _sample_results()
    csv_path = tmp_path / "results.csv"
    xlsx_path = tmp_path / "results.xlsx"
    json_path = tmp_path / "results.json"

    export_results_csv(results, str(csv_path))
    export_results_xlsx(results, str(xlsx_path))
    export_results_json(results, str(json_path))

    with csv_path.open(encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream))
    assert "display version" in rows[1]
    assert "H3C Comware Software" in rows[1]

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["执行摘要", "命令明细"]
        assert workbook["执行摘要"]["A2"].value == "SW1"
        assert workbook["执行摘要"]["K2"].value == "1.2"
        assert workbook["执行摘要"]["L2"].value == "0.8"
        assert workbook["命令明细"]["D2"].value == "display version"
    finally:
        workbook.close()

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["results"][0]["model_detected"] == "S6520"


def test_result_dialog_has_maximize_without_unused_help_button():
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    from PyQt5.QtCore import Qt
    from PyQt5.QtWidgets import QApplication
    from ui.result_dialog import ResultCenterDialog

    app = QApplication.instance() or QApplication([])
    dialog = ResultCenterDialog(_sample_results())
    try:
        flags = dialog.windowFlags()
        assert flags & Qt.WindowMaximizeButtonHint
        assert not flags & Qt.WindowContextHelpButtonHint
        assert dialog.table.horizontalHeaderItem(6).text() == "总耗时"
        assert dialog.table.horizontalHeaderItem(7).text() == "连接准备"
        assert dialog.table.horizontalHeaderItem(8).text() == "任务执行"
        assert dialog.table.item(0, 7).text() == "1.20 秒"
        assert dialog.table.item(0, 8).text() == "0.80 秒"
    finally:
        dialog.close()
        app.processEvents()


def test_connection_result_splits_total_connection_and_operation_time():
    device = DeviceInfo(
        "h3c", "192.0.2.10", 22, "admin", "secret", "SW1"
    )
    connection = SSHConnection(device)
    connection.started_at = datetime.now() - timedelta(seconds=5)
    connection.finished_at = connection.started_at + timedelta(seconds=5)
    connection.connection_duration_seconds = 3.2

    result = connection.get_connection_info()

    assert result["duration_seconds"] == 5.0
    assert result["connection_duration_seconds"] == 3.2
    assert result["operation_duration_seconds"] == 1.8


class _H3CTestServer(paramiko.ServerInterface):
    def check_auth_password(self, username, password):
        if username == "admin" and password == "secret":
            return paramiko.AUTH_SUCCESSFUL
        return paramiko.AUTH_FAILED

    def check_channel_request(self, kind, _channel_id):
        return (
            paramiko.OPEN_SUCCEEDED
            if kind == "session"
            else paramiko.OPEN_FAILED_ADMINISTRATIVELY_PROHIBITED
        )

    @staticmethod
    def check_channel_pty_request(*_args):
        return True

    @staticmethod
    def check_channel_shell_request(_channel):
        return True


def _serve_h3c_once(listener, host_key, errors):
    transport = channel = None
    try:
        client_socket, _ = listener.accept()
        transport = paramiko.Transport(client_socket)
        transport.add_server_key(host_key)
        transport.start_server(server=_H3CTestServer())
        channel = transport.accept(5)
        if channel is None:
            raise RuntimeError("SSH test channel was not opened")
        channel.send("\r\n<H3C>")
        buffer = ""
        while transport.is_active():
            data = channel.recv(4096)
            if not data:
                break
            buffer += data.decode("utf-8", errors="ignore")
            while "\n" in buffer:
                command, buffer = buffer.split("\n", 1)
                command = command.strip()
                if not command:
                    continue
                if command == "display version":
                    output = (
                        "\r\nH3C S6520 uptime is 1 week\r\n"
                        "H3C Comware Software, Version 7\r\n<H3C>"
                    )
                elif command == "display clock":
                    output = "\r\n12:00:00 UTC Sun 07/27/2026\r\n<H3C>"
                else:
                    output = f"\r\n{command}\r\n<H3C>"
                channel.send(output)
    except Exception as exc:
        errors.append(exc)
    finally:
        if channel is not None:
            try:
                channel.close()
            except (EOFError, OSError):
                pass
        if transport is not None:
            try:
                transport.close()
            except (EOFError, OSError):
                pass
        listener.close()


def test_real_paramiko_handshake_tofu_and_h3c_shell(monkeypatch, tmp_path):
    listener = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    listener.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    listener.bind(("127.0.0.1", 0))
    listener.listen(1)
    port = listener.getsockname()[1]
    errors = []
    thread = threading.Thread(
        target=_serve_h3c_once,
        args=(listener, paramiko.RSAKey.generate(1024), errors),
        daemon=True,
    )
    thread.start()
    known_hosts = tmp_path / "known_hosts"
    monkeypatch.setenv("AOMT_KNOWN_HOSTS_PATH", str(known_hosts))

    device = DeviceInfo(
        "h3c", "127.0.0.1", port, "admin", "secret", "H3C-SIM",
        host_key_policy="tofu",
    )
    connection = SSHConnection(device)
    assert connection.connect() is True
    assert connection.brand_detected == "h3c"
    output = connection.execute_command("display clock")
    connection.disconnect()
    thread.join(timeout=5)

    assert "12:00:00" in output
    assert known_hosts.is_file()
    assert known_hosts.read_text(encoding="utf-8").strip()
    assert errors == []
