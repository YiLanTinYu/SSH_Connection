import sys
import json
import threading
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from config.device_config import DeviceConfigManager, DeviceInfo
from config.device_commands import detect_brand
from config.builtin_templates import (
    BUILTIN_TEMPLATE_DEFINITIONS,
    get_builtin_templates,
)
from config.template_renderer import render_template
from core.ssh_manager_simple import SSHConnection, SSHManager
from utils.password_crypto import PasswordDecryptionError, decrypt_password, encrypt_password


def write_excel(path, rows):
    workbook = Workbook()
    sheet = workbook.active
    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    workbook.save(path)


def test_excel_import_rejects_blank_required_cells(tmp_path):
    path = tmp_path / "devices.xlsx"
    write_excel(path, [
        {"ip": None, "username": "admin", "password": "pwd", "brand": "h3c", "port": 22},
        {"ip": "192.168.1.1", "username": None, "password": "pwd", "brand": "h3c", "port": 22},
        {"ip": "192.168.1.2", "username": "admin", "password": None, "brand": "h3c", "port": 22},
        {"ip": "192.168.1.3", "username": "admin", "password": "pwd", "brand": None, "port": None},
    ])

    mgr = DeviceConfigManager()
    ok, fail, errors = mgr.import_from_excel(str(path))

    assert ok == 1
    assert fail == 3
    assert len(errors) == 3
    device = mgr.get_devices()[0]
    assert device.brand == "h3c"
    assert device.port == 22
    assert device.ip == "192.168.1.3"


def test_excel_import_validates_ip_and_port(tmp_path):
    path = tmp_path / "bad_devices.xlsx"
    write_excel(path, [
        {"ip": "999.1.1.1", "username": "admin", "password": "pwd", "brand": "h3c", "port": 22},
        {"ip": "192.168.1.1", "username": "admin", "password": "pwd", "brand": "h3c", "port": 70000},
    ])

    mgr = DeviceConfigManager()
    ok, fail, errors = mgr.import_from_excel(str(path))

    assert ok == 0
    assert fail == 2
    assert any("IP" in e for e in errors)
    assert any("port" in e for e in errors)


def test_excel_import_skips_duplicate_ip_port(tmp_path):
    path = tmp_path / "duplicate_devices.xlsx"
    write_excel(path, [
        {"ip": "192.168.1.1", "username": "admin", "password": "pwd", "brand": "h3c", "port": 22},
        {"ip": "192.168.1.1", "username": "admin2", "password": "pwd2", "brand": "huawei", "port": 22},
        {"ip": "192.168.1.1", "username": "admin3", "password": "pwd3", "brand": "h3c", "port": 2222},
    ])

    mgr = DeviceConfigManager()
    ok, fail, errors = mgr.import_from_excel(str(path))

    assert ok == 2
    assert fail == 0
    assert errors == []
    assert mgr.last_import_skipped_count == 1
    assert len(mgr.get_devices()) == 2


def test_per_device_script_resolves_only_by_device_name(tmp_path):
    for name in ("SW1.txt", "192.168.1.1.txt", "h3c.txt", "default.txt"):
        (tmp_path / name).write_text("display version\n", encoding="utf-8")
    manager = SSHManager()
    manager.command_directory = str(tmp_path)

    by_name = DeviceInfo("h3c", "192.168.1.1", name="SW1")
    no_name_match = DeviceInfo("h3c", "192.168.1.1", name="OTHER")

    assert Path(manager.resolve_command_file(by_name)).name == "SW1.txt"
    assert manager.resolve_command_file(no_name_match) is None


def test_device_name_is_sanitized_for_windows_filename(tmp_path):
    (tmp_path / "SW_核心.txt").write_text("display version\n", encoding="utf-8")
    manager = SSHManager()
    manager.command_directory = str(tmp_path)
    device = DeviceInfo("h3c", "2026:1000:120::23", name="SW:核心")

    assert Path(manager.resolve_command_file(device)).name == "SW_核心.txt"


def test_add_device_rejects_duplicate_ip_port():
    mgr = DeviceConfigManager()

    assert mgr.add_device(DeviceInfo("h3c", "192.168.1.1", 22, "admin", "pwd"))
    assert not mgr.add_device(DeviceInfo("huawei", "192.168.1.1", 22, "admin2", "pwd2"))
    assert mgr.add_device(DeviceInfo("h3c", "192.168.1.1", 2222, "admin3", "pwd3"))
    assert len(mgr.get_devices()) == 2


def test_json_load_validates_records_and_skips_duplicates(tmp_path):
    path = tmp_path / "devices.json"
    path.write_text(json.dumps([
        {"ip": "999.1.1.1", "username": "admin", "password": "pwd", "brand": "h3c", "port": 22},
        {"ip": "192.168.1.1", "username": "admin", "password": "pwd", "brand": "h3c", "port": 22},
        {"ip": "192.168.1.1", "username": "admin2", "password": "pwd2", "brand": "huawei", "port": 22},
        {"ip": "192.168.1.2", "username": "admin", "brand": "h3c", "port": 22},
    ]), encoding="utf-8")

    mgr = DeviceConfigManager()

    assert mgr.load_from_json(str(path))
    assert len(mgr.get_devices()) == 1
    assert mgr.get_devices()[0].ip == "192.168.1.1"
    assert mgr.last_import_skipped_count == 1


def test_json_load_rejects_non_list_without_clearing_existing_devices(tmp_path):
    path = tmp_path / "bad.json"
    path.write_text(json.dumps({"ip": "192.168.1.1"}), encoding="utf-8")

    mgr = DeviceConfigManager()
    mgr.add_device(DeviceInfo("h3c", "192.168.1.1", 22, "admin", "pwd"))

    assert not mgr.load_from_json(str(path))
    assert len(mgr.get_devices()) == 1


def test_excel_export_omits_passwords_by_default(tmp_path):
    path = tmp_path / "export.xlsx"
    mgr = DeviceConfigManager()
    mgr.add_device(DeviceInfo("h3c", "192.168.1.1", 22, "admin", "secret"))

    assert mgr.export_to_excel(str(path))

    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    password_col = headers.index("password") + 1
    assert sheet.cell(row=2, column=password_col).value in (None, "")


def test_excel_export_encrypts_passwords_when_master_password_is_given(tmp_path):
    path = tmp_path / "export_with_password.xlsx"
    mgr = DeviceConfigManager()
    mgr.add_device(DeviceInfo("h3c", "192.168.1.1", 22, "admin", "secret"))

    assert mgr.export_to_excel(str(path), include_password=True, master_password="MasterPass123")

    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    password_col = headers.index("password") + 1
    encrypted = sheet.cell(row=2, column=password_col).value
    assert encrypted.startswith("AOMT_ENC_V1$")
    assert "secret" not in encrypted
    assert decrypt_password(encrypted, "MasterPass123") == "secret"


def test_password_cipher_rejects_wrong_master_password():
    encrypted = encrypt_password("SwitchSecret!", "CorrectMaster123")

    with pytest.raises(PasswordDecryptionError):
        decrypt_password(encrypted, "WrongMaster123")


def test_encrypt_excel_copy_and_import_with_master_password(tmp_path):
    source = tmp_path / "plain.xlsx"
    target = tmp_path / "encrypted.xlsx"
    write_excel(source, [{
        "name": "SW1", "brand": "h3c", "ip": "192.168.1.1", "port": 22,
        "username": "admin", "password": "DeviceSecret!",
    }])

    assert DeviceConfigManager.inspect_excel_password_mode(str(source)) == "plain"
    assert DeviceConfigManager.encrypt_excel_passwords(
        str(source), str(target), "MasterPass123"
    ) == 1
    assert DeviceConfigManager.inspect_excel_password_mode(str(target)) == "encrypted"

    mgr = DeviceConfigManager()
    ok, fail, errors = mgr.import_from_excel(str(target), master_password="MasterPass123")
    assert (ok, fail, errors) == (1, 0, [])
    assert mgr.get_devices()[0].password == "DeviceSecret!"


def test_unknown_brand_stays_unknown_until_user_fallback():
    assert detect_brand("") == "unknown"
    assert detect_brand("% Invalid input detected") == "unknown"


def test_ssh_ipv4_extraction_ignores_invalid_addresses():
    connection = SSHConnection(DeviceInfo("h3c", "192.0.2.1", 22, "admin", "pwd"))

    assert connection._extract_ipv4("bad 999.999.999.999\nnext 192.168.1.1") == "192.168.1.1"
    assert connection._extract_ipv4("only 999.999.999.999") == ""


def test_connection_result_keeps_task_success_after_disconnect():
    connection = SSHConnection(DeviceInfo("h3c", "192.0.2.1", 22, "admin", "pwd"))
    connection.is_connected = True
    connection.task_success = True

    connection.disconnect()
    result = connection.get_connection_info()

    assert result["is_connected"] is True
    assert result["task_success"] is True
    assert result["ssh_active"] is False


class ConfirmShell:
    def __init__(self):
        self.sent = []
        self.outputs = ["Are you sure? [Y/N]", "<H3C>"]

    def send(self, text):
        self.sent.append(text.strip())

    def recv_ready(self):
        return bool(self.outputs)

    def recv(self, _size):
        return self.outputs.pop(0).encode("utf-8")


def test_save_config_confirms_interactive_prompt(monkeypatch):
    import config.device_commands as commands

    monkeypatch.setattr(commands, "get_command", lambda _brand, _key: "save")
    connection = SSHConnection(DeviceInfo("h3c", "192.0.2.1", 22, "admin", "pwd"))
    connection.is_connected = True
    connection.brand_detected = "h3c"
    connection._shell = ConfirmShell()

    assert connection.save_config()
    assert connection._shell.sent == ["save", "y"]


def test_cancelled_connection_stops_command_loop():
    cancel_event = threading.Event()
    cancel_event.set()
    connection = SSHConnection(
        DeviceInfo("h3c", "192.0.2.1", 22, "admin", "pwd"),
        cancel_event=cancel_event,
    )

    assert connection.execute_commands(["display version"]) == []
    assert connection.execute_command("display version") == "任务已取消"


class ChunkedPromptShell:
    def __init__(self):
        self.outputs = [
            b"CPU Usage : 9%\\r\\nVIDL 91%",
            b"\\r\\n<HUAWEI>",
        ]

    def recv_ready(self):
        return bool(self.outputs)

    def recv(self, _size):
        return self.outputs.pop(0)

    def send(self, _text):
        return None


def test_prompt_detection_does_not_stop_at_percentage_output():
    connection = SSHConnection(
        DeviceInfo("huawei", "192.0.2.1", 22, "admin", "pwd")
    )
    connection._shell = ChunkedPromptShell()

    output = connection._read_until_prompt(timeout=0.2)

    assert "VIDL 91%" in output
    assert output.rstrip().endswith("<HUAWEI>")
    assert connection._shell.outputs == []
    assert not connection._has_terminal_prompt("CPU Usage : 91%")
    assert connection._has_terminal_prompt("<HUAWEI>")
    assert connection._has_terminal_prompt("[HUAWEI]")
    assert connection._has_terminal_prompt("Router#")


def test_business_commands_are_sent_without_brand_translation(monkeypatch):
    connection = SSHConnection(DeviceInfo("cisco", "192.0.2.1", 22, "admin", "pwd"))
    connection.brand_detected = "cisco"
    commands = ["display version", "vlan 10", "show running-config"]
    sent = []

    def execute_literal(command):
        sent.append(command)
        return f"output: {command}"

    monkeypatch.setattr(connection, "execute_command", execute_literal)
    results = connection.execute_commands(commands)

    assert sent == commands
    assert [item["command"] for item in results] == commands


def test_builtin_templates_are_parameterized_brand_specific_configurations():
    templates = get_builtin_templates()
    assert len(templates) == len(BUILTIN_TEMPLATE_DEFINITIONS)
    assert len(templates) == 14
    assert {item["brand"] for item in templates} == {"h3c", "huawei"}
    for template in templates:
        path = Path(template["path"])
        assert path.is_file()
        values = {
            field["name"]: (
                "AomtPass!2026"
                if field.get("sensitive")
                else field.get("default", "")
            )
            for field in template["parameters"]
        }
        rendered = render_template(template, values)
        assert rendered.commands
        assert "{{" not in "\n".join(rendered.commands)
        assert template["brand"] in ("h3c", "huawei")
